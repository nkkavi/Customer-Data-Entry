import tkinter
from tkinter import messagebox
from tkinter import *
from tkinter import ttk
import customtkinter
import pandas as pd
from CTkMessagebox import CTkMessagebox
from tkcalendar import DateEntry
from PIL import Image
import openpyxl
import os
import datetime
import webbrowser
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg

## code for py to exe
## pyinstaller --noconfirm --onefile --windowed --add-data "C:\Users\raju8\Music\customtkinter;customtkinter" --add-data "C:\Users\raju8\Music\\CTkMessagebox;\CTkMessagebox" --add-data "nk32.png;." --add-data "icon.ico;." --hidden-import babel.numbers --icon=r64.png main6.py


## Defining the functions

file1 = "data.xlsx"
filepath_1 = file1


## file location creating and finding


def load_data():
    print("loaded from internal storage")
    if not os.path.exists(filepath_1):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        heading = ["name", "mobile", "cus_labour", "payment", "paid", "due", "channel", "area", "day"]
        sheet.append(heading)
        workbook.save(filepath_1)
        print("New File created")
    workbook = openpyxl.load_workbook(filepath_1)
    sheet = workbook.active
    workbook.save(filepath_1)

    list_values = list(sheet.values)
    print(list_values)
    for col_name in cols:
        treeview.heading(col_name, text=col_name)
    for value_tuple in list_values[1:]:
        treeview.insert('', customtkinter.END, values=value_tuple)


## Clearing the text from username and all fields


def clear_added():
    combobox_var_cus_lab.set("Customer")
    combobox_var_Zone.set("Area")
    combobox_var_method.set("Cash")
    combobox_var_Package.set("F+Tel_SP")
    username.delete(0, tkinter.END)
    con_number.delete(0, END)
    Money.delete(0, END)
    Money_due.delete(0, END)
    print("Successfully data cleared from the text Add_clear succeed")


def clear_search():
    cus_search.delete(0, END)
    clear_added()
    print('search clear')


# Exit Button
def exit_button():
    result = messagebox.askquestion("Exit", "Are you sure you want to exit?")
    if result == "yes":
        clear_search()
        clear_added()
        app.destroy()
        print('exit work')


## Customer Payment Paid / Unpaid


def combobox_cus_lab(choice):
    print("combobox dropdown clicked:", choice)

## Transition Mode


def combobox_method(choice):
    print("combobox dropdown clicked:", choice)

## Channel Package


def combobox_package(choice):
    print("Combobox dropbox clicked:", choice)

## Location of Customer


def combobox_zone(choice):
    print("combobox dropdown clicked:", choice)


def is_numeric(s):
    try:
        float(s)
        return True
    except ValueError:
        return False


def add_button():
    name = username.get()
    if name:
        paid = Money.get()
        if not is_numeric(paid) or float(paid) < 0:
            CTkMessagebox(title="Error", message="Invalid paid amount! Please enter a non-negative number.",
                          option_1="Retry")
            print("Error1")
        else:
            due = Money_due.get()
            if not is_numeric(due):
                CTkMessagebox(title="Error",
                              message="Invalid due amount! \nIf Customer paid Due means \nkeep negative sign Due amount"
                                      "\n If no Due means Entry 0\n If Due means Entry Due amount",
                              option_1="Retry")
                print("Error2")
            else:
                mobile = con_number.get()
                cus_labour = combobox_cus_lab.get()
                payment = combobox_Method1.get()
                channel = combobox_Package1.get()
                area = combobox_Zone1.get()
                day = Date1.get()
                clear_added()
                print("user name:", name, "mobile number:", mobile, "Cus/Labour:", cus_labour)
                print("payment method:", payment, "paid amount:", paid, "due amount", due)
                print("Channel Package:", channel, "Location:", area, "Date:", day)

                if not os.path.exists(filepath_1):
                    workbook = openpyxl.Workbook()
                    sheet = workbook.active
                    heading = ["name", "mobile", "cus_labour", "payment", "paid", "due", "channel", "area", "day"]
                    sheet.append(heading)
                    workbook.save(filepath_1)
                workbook = openpyxl.load_workbook(filepath_1)
                sheet = workbook.active
                row_values = [name, mobile, cus_labour, payment, paid, due, channel, area, day]
                sheet.append(row_values)
                workbook.save(filepath_1)
                treeview.insert('', tkinter.END, values=row_values)
                print("22222 File was created in internal storage")

                CTkMessagebox(title="Customer details added to list", message="Successfully!!!", icon="check",
                              option_1="Thanks")

    else:
        CTkMessagebox(title="Error", message="Username are required!!!", option_1="Retry")
        print("Error3")


## delete Customer Details

def delete_button():
    result = messagebox.askquestion("Delete", "Are you sure you want to Delete?")
    if result == "yes":
        selected_item = treeview.selection()
        if not selected_item:
            return
        selected_item = selected_item[0]
        row_index = int(treeview.index(selected_item))

        workbook = openpyxl.load_workbook(filepath_1)
        sheet = workbook.active
        sheet.delete_rows(row_index + 2, amount=1)
        workbook.save(filepath_1)
        treeview.delete(selected_item)

        CTkMessagebox(title="Customer data deleted", message="Successfully Deleted Record!!!", icon="check",
                      option_1="Done")


def resource_path(relative_path):
    try:
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")

    return os.path.join(base_path, relative_path)


def open_web_link():
    webbrowser.open_new("https://lcoportal.nxtdigital.in/login.php")


app = customtkinter.CTk()
# create a photo image object from the PNG file
logo_app = resource_path("icon.ico")
app.iconbitmap(logo_app)
app.geometry("697x538")
app.title("karupothula Sarangapani")


customtkinter.set_default_color_theme("green")
customtkinter.set_appearance_mode("dark")


frame = customtkinter.CTkFrame(master=app)
frame.grid()

notebook = customtkinter.CTkTabview(frame)
notebook.pack(fill='both', expand=True)


notebook.add("Customers Details")
notebook.add("Graphs")
notebook.add("Contact")
notebook.set("Customers Details")

## customer Details & Scroll Bar
## Scroll Bar


treeFrame = customtkinter.CTkFrame(master=notebook.tab("Customers Details"))
treeFrame.grid(row=0, column=0, padx=10)

y_scrollbar = customtkinter.CTkScrollbar(master=treeFrame, width=12, height=6, fg_color="transparent")
y_scrollbar.pack(side="right", fill="y")


cols = ("name", "mobile", "cus_labour", "payment", "paid", "due", "channel", "area", "day")

treeview = ttk.Treeview(treeFrame, show="headings", yscrollcommand=y_scrollbar.set, selectmode="browse",
                        columns=cols, height=10)

treeview.column("name", width=100)
treeview.column("mobile", width=80)
treeview.column("cus_labour", width=60)
treeview.column("payment", width=50)
treeview.column("paid", width=50)
treeview.column("due", width=50)
treeview.column("channel", width=60)
treeview.column("area", width=90)
treeview.column("day", width=70)
y_scrollbar.configure(command=treeview.yview)

treeview.pack(ipadx=21, ipady=10)
load_data()


tab1label1 = customtkinter.CTkLabel(master=notebook.tab("Customers Details"), text=" ", fg_color=("#333333", "#333333"))
tab1label1.grid(row=1, column=0, padx=1, pady=5)

Cus_frame = customtkinter.CTkLabel(master=tab1label1, text="ॐ",
                                   text_color=("Red", "white"), font=("Bold", 19))
Cus_frame.grid(row=0, column=0)


def get_value(event):
    clear_added()
    row_id = treeview.selection()[0]
    select = treeview.set(row_id)
    username.insert(0, select['name'])
    con_number.insert(0, select['mobile'])
    Money.insert(0, select['paid'])
    Money_due.insert(0, select['due'])

    # Get the selected value from the Treeview and insert it into the Combobox
    selected_cus_lab = select['cus_labour']
    combobox_cus_lab.set(selected_cus_lab)

    # Get the selected value from the Treeview and insert it into the Combobox
    selected_payment = select['payment']
    combobox_Method1.set(selected_payment)

    # Get the selected value from the Treeview and insert it into the Combobox
    selected_channel = select['channel']
    combobox_Package1.set(selected_channel)

    # Get the selected value from the Treeview and insert it into the Combobox
    selected_area = select['area']
    combobox_Zone1.set(selected_area)


df = pd.read_excel(filepath_1)

def my_search():
    if len(treeview.get_children()) > 0:
        treeview.delete(*treeview.get_children())

    l1 = list(df)  # List of column names as headers
    query = cus_search.get().strip()  # get user entered string
    if query.isdigit():  # if query is number
        str1 = df["mobile"] == int(query)  #
    else:
        str1 = df.name.str.contains(query, case=False)  # name column value matching
    df2 = df[(str1)]   # combine all conditions using | operator
    r_set = df2.to_numpy().tolist()  # Create list of list using rows
    clear_search()
    treeview.bind('<Double-Button-1>', get_value)

    for i in l1:
        if i in treeview["columns"]:
            treeview.column(i, width=75, anchor="c")
            # Headings of respective columns
            treeview.heading(i, text=i)
    for dt in r_set:
        v = [r for r in dt]  # creating a list from each row
        treeview.insert("", "end", values=v)  # adding row




current_theme = 'dark'


def switch_theme():
    global current_theme
    if current_theme == 'light':
        customtkinter.set_appearance_mode('dark')
        current_theme = 'dark'
        switch_mode.configure(text='Dark Mode')
    else:
        customtkinter.set_appearance_mode('light')
        current_theme = 'light'
        switch_mode.configure(text='Light Mode')


cus_search = customtkinter.CTkEntry(master=tab1label1, placeholder_text="Search: Name/Mobile No.", width=180)
cus_search.grid(row=0, column=1)

search_button = customtkinter.CTkButton(master=tab1label1, text="Find", command=lambda: my_search(), width=50)
search_button.grid(row=0, column=2)


switch_mode_var = customtkinter.StringVar()
switch_mode = customtkinter.CTkSwitch(master=tab1label1, text='Mode', text_color="white", variable=switch_mode_var,
                                      command=switch_theme)
switch_mode.grid(row=0, column=3, pady=10)


username = customtkinter.CTkEntry(master=tab1label1, placeholder_text="User Name")
username.grid(row=1, column=0)

con_number = customtkinter.CTkEntry(master=tab1label1, placeholder_text="Mobile Number")
con_number.grid(row=1, column=1)

combobox_var_cus_lab = customtkinter.StringVar(value="Customer")  # set initial value
combobox_cus_lab = customtkinter.CTkComboBox(master=tab1label1, values=[" ", "Customer", "Labour"],
                                             command=combobox_cus_lab,
                                             variable=combobox_var_cus_lab)
combobox_cus_lab.grid(row=1, column=2)

combobox_var_method = customtkinter.StringVar(value="Cash")  # set initial value
combobox_Method1 = customtkinter.CTkComboBox(master=tab1label1,
                                             values=[" ", "Cash", "GPay", "PhonePay", "PayTM", "Paid", "Due"],
                                             command=combobox_method, variable=combobox_var_method)
combobox_Method1.grid(row=2, column=0)

Money = customtkinter.CTkEntry(master=tab1label1, placeholder_text="Paid Amount")
Money.grid(row=2, column=1)

Money_due = customtkinter.CTkEntry(master=tab1label1, placeholder_text="Due Amount")
Money_due.grid(row=2, column=2)
combobox_var_Package = customtkinter.StringVar(value="F+Tel_SP")  # set initial value
combobox_Package1 = customtkinter.CTkComboBox(master=tab1label1, values=[" ", "F+Tel_SP", "F+Tel_DP"],
                                              command=combobox_package,
                                              variable=combobox_var_Package)
combobox_Package1.grid(row=3, column=0)

combobox_var_Zone = customtkinter.StringVar(value="Pusaverla Palli")  # set initial value
combobox_Zone1 = customtkinter.CTkComboBox(master=tab1label1,
                                           values=[" ", "Pusaverla Palli", "Rajeswar Rao", "Ldpet MainRoad", "Ldpet Thand", "Ldpet Old"],
                                           command=combobox_zone,
                                           variable=combobox_var_Zone)
combobox_Zone1.grid(row=3, column=1)

Date1 = DateEntry(master=tab1label1, date_pattern='dd-mm-yyyy')
Date1.grid(row=3, column=2, sticky='w')

for widget in tab1label1.winfo_children():
    widget.grid_configure(padx=15, pady=9)


## Update Section

tab1label2 = customtkinter.CTkLabel(master=notebook.tab("Customers Details"), text=" ",
                                    fg_color=("#333333", "#333333"))
tab1label2.grid(row=3, column=0, padx=0, pady=0)


nxt_button = customtkinter.CTkButton(master=tab1label2, text="Open NXT", command=open_web_link,
                                        text_color=("white", "white"), fg_color=("#2e8dba", "#2e8dba"),
                                        hover_color=("#5ec46f", "#5ec46f"), width=100)
nxt_button.grid(row=0, column=0)

add_button = customtkinter.CTkButton(master=tab1label2, text="Add Details", command=add_button, width=100)
add_button.grid(row=0, column=1)

delete_button = customtkinter.CTkButton(master=tab1label2, text="Delete Details", command=delete_button,
                                        text_color=("white", "white"), fg_color=("red", "red"),
                                        hover_color=("#2e8dba", "#2e8dba"), width=100)
delete_button.grid(row=0, column=2)

close_button = customtkinter.CTkButton(master=tab1label2, text="EXIT", text_color=("white", "white"),
                                       fg_color=("#cc43d8", "#cc43d8"), hover_color=("#ef2d30", "#ef2d30"),
                                       command=exit_button, width=50)
close_button.grid(row=0, column=3)

for widget in tab1label2.winfo_children():
    widget.grid_configure(padx=15, pady=11)

## Tab 2

def due_customers():
    top = customtkinter.CTkToplevel(notebook.tab("Graphs"))
    # create a photo image object from the PNG file
    logo_top = resource_path("icon.ico")
    #top.iconbitmap(logo_top)
    top.iconbitmap(False, logo_top)

    top.title("Due Customers Data")

    def load_due_data():
        ## global customer_sales

        # Set the date column as the index
        ## df.set_index('day', inplace=True)
        # Extract the month and year from the date column
        df['month'] = pd.DatetimeIndex(df.index).month
        df['year'] = pd.DatetimeIndex(df.index).year
        # Group the data by customer and get the sum of sales
        customer_sales = df.groupby(['name', 'area', 'mobile'])['due'].sum().reset_index()
        # Sort the data by payment in descending order
        customer_sales = customer_sales.sort_values(by='due', ascending=False)
        # Display the report in a text box
        customer_sales_string = customer_sales[['due', 'name', 'mobile', 'area']].to_string(index=False, col_space=35)
        text_box.delete('1.0', "end")
        text_box.insert("end", customer_sales_string)
        clear_due_search()

    # Create a function to search for names
    def search_due_names():
        # Get the search term from the entry box
        search_term = search_due_entry.get()
        customer_sales = df.groupby(['name', 'area', 'mobile'])['due'].sum().reset_index()

        # Filter the customer sales data by the search term
        result = customer_sales[customer_sales['name'].str.contains(search_term, case=False)]
        # Sort the data by payment in descending order
        result = result.sort_values(by='due', ascending=False)
        # Display the search results in the text box
        customer_sales_string = result[['due', 'name', 'mobile', 'area']].to_string(index=False, col_space=35)
        text_box.delete('1.0', "end")
        text_box.insert("end", customer_sales_string)
        #clear_due_search()

    # Create a button to load the Excel file
    load_due_button = customtkinter.CTkButton(top, text="Load Due Customers Data", command=load_due_data)
    load_due_button.grid(row=0, column=0)

    # Create a label and entry box for searching names
    def clear_due_search():
        search_due_entry.delete(0, END)
        print('search clear from due customers data entry')

    search_due_entry = customtkinter.CTkEntry(top, placeholder_text="search Due Names")
    search_due_entry.grid(row=1, column=0)
    # Create a button to search for names
    search_due_button = customtkinter.CTkButton(top, text="Search", command=search_due_names)
    search_due_button.grid(row=2, column=0)

    # Create a text box to display the report
    text_box = customtkinter.CTkTextbox(top, width=680, height=390)
    text_box.grid(row=3, column=0)
    import os
    from reportlab.pdfgen import canvas
    from reportlab.lib.pagesizes import letter
    from datetime import datetime

    def print_pdf(search_term=None):
        if search_term:
            # Filter the customer sales data by the search term
            result = df[df['name'].str.contains(search_term, case=False)]
            customer_sales = result.groupby(['name', 'area', 'mobile'])['due'].sum().reset_index()
        else:
            # Get the customer list from the text box
            customer_sales = df.groupby(['name', 'area', 'mobile'])['due'].sum().reset_index()

        customer_sales = customer_sales.sort_values(by='due', ascending=False)
        customer_sales_string = customer_sales[['due', 'name', 'mobile', 'area']].to_string(index=False, col_space=35)

        # Create a PDF file with timestamp
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        pdf_file_name = f"customer_list_{timestamp}.pdf"
        c = canvas.Canvas(pdf_file_name, pagesize=letter)
        textobject = c.beginText()
        textobject.setTextOrigin(50, 750)
        textobject.setFont("Helvetica", 12)
        textobject.textLines(customer_sales_string)
        c.drawText(textobject)
        c.save()

        # Open the PDF file
        os.system('start ' + pdf_file_name)

    # Create a button to print the customer list in PDF format
    print_pdf_button = customtkinter.CTkButton(top, text="Print PDF", command=lambda: print_pdf(search_due_entry.get()))


    print_pdf_button.grid(row=2, column=0, sticky="e")



    for widget in top.winfo_children():
        widget.grid_configure(padx=15, pady=9)


tab2label0 = customtkinter.CTkLabel(master=notebook.tab("Graphs"), text=" ",
                                    fg_color=("#333333", "#333333"))
tab2label0.grid(row=0, column=0, padx=1, pady=1)

Imp_frame = customtkinter.CTkLabel(master=tab2label0, text="ॐ",
                                   text_color=("Red", "white"), font=("Bold", 19))
Imp_frame.grid(row=0, column=0)

browser_nxt_button = customtkinter.CTkButton(master=tab2label0, text="Open NXT",
                                             command=open_web_link, width=20)
browser_nxt_button.grid(row=0, column=1, sticky="e")


due_button = customtkinter.CTkButton(tab2label0, text="Due Amount Customers", command=due_customers, width=20)
due_button.grid(row=0, column=1, sticky="w")

# Define function to create and show the pie chart
paid_sum = df['paid'].sum()
due_sum = df['due'].sum()

# Check if there is any data available
if paid_sum == 0 and due_sum == 0:
    # Display a message if there is no data available
    label_check1 = customtkinter.CTkLabel(tab2label0, text='No data available to display donut pie',
                                          text_color=('#e5ef56', '#e5ef56'), font=('Arial', 12))
    label_check1.grid(row=1, column=0, padx=10, pady=10)

    # If there is no data available, skip the pie chart
    pass
elif paid_sum + due_sum <= 0:
    # Display a message if there is no positive data available
    label_check2 = customtkinter.CTkLabel(tab2label0, text='No data available to display donut pie',
                                          text_color=('#e5ef56', '#e5ef56'), font=('Arial', 12))
    label_check2.grid(row=1, column=0, padx=10, pady=10)

    # If there is no positive data available, skip the pie chart
    pass
else:
    # If there is data available, create the pie chart
    labels = ['Paid', 'Due']
    sizes = [paid_sum, due_sum]
    colors = ['#1f77b4', '#ff7f0e']
    # Check if any size is negative
    if any(x < 0 for x in sizes):
        # Display a message if there are negative sizes
        label_check3 = customtkinter.CTkLabel(tab2label0, text='Negative values are not allowed \nto display donut pie',
                                              text_color=('#e5ef56', '#e5ef56'), font=('Arial', 12))
        label_check3.grid(row=1, column=0, padx=10, pady=10)

        # If there are negative sizes, skip the pie chart
        pass
    else:
        # Create the pie chart
        circle = plt.Circle((0, 0), 0.7, color='lightgrey')
        fig, ax = plt.subplots(figsize=(2, 2), dpi=100)
        fig.patch.set_facecolor('#f497eb')
        ax.pie(sizes, colors=colors, startangle=90, counterclock=False, radius=1.0,
               wedgeprops=dict(width=0.3, edgecolor='white'), labels=labels, autopct='%1.1f%%', pctdistance=0.85)
        ax.add_artist(circle)
        ax.axis('equal')

        # Add a legend to the chart
        legend = ax.legend(loc='center', bbox_to_anchor=(-0.5, 0.5))
        legend.set_title('Labels', prop={'size': 10})

        # Add the chart to the window
        canvas = FigureCanvasTkAgg(fig, master=tab2label0)
        canvas.draw()
        canvas.get_tk_widget().grid(row=1, column=0, padx=0, pady=0, sticky="nsew")

## Area Due Amount
import numpy as np


def load_due_graph():
    # Set the date column as the index
    df.set_index('day', inplace=True)
    # Extract the month and year from the date column
    df['month'] = pd.DatetimeIndex(df.index).month
    df['year'] = pd.DatetimeIndex(df.index).year
    # Group the data by customer and get the sum of sales
    customer_sales1 = df.groupby(['area'])['due'].sum().reset_index()
    # Sort the data by payment in descending order
    customer_sales1 = customer_sales1.sort_values(by='due', ascending=False)
    # Return the top 5 customers by due amount
    return customer_sales1.head(5)

# Load data
customer_sales = load_due_graph()

# Check if there is any data available
if not customer_sales.empty:
    # Define colors for bars
    colors = ['#1f77b4', '#ff7f0e', '#2ca02c', '#d62728', '#9467bd']

    # Create bar chart
    fig, ax = plt.subplots(figsize=(4, 2))
    rects = ax.bar(customer_sales['area'], customer_sales['due'], color=colors[:5])
    #ax.set_xlabel('Location Names', fontsize=9)
    #ax.set_ylabel('Due Amount', fontsize=9)
    ax.set_title('Top 5 Area by Due Amount', fontsize=10)
    plt.setp(ax.get_xticklabels(), rotation=0, ha="right", fontsize=8)
    ax.set_xticklabels([])
    plt.tight_layout()
    ax.set_yticklabels([])
    plt.tight_layout()
    #plt.show()



    # Add x label names on the bar inside the bar
    for rect, (area, due) in zip(rects, zip(customer_sales['area'], customer_sales['due'])):
        height = rect.get_height()
        ax.text(rect.get_x() + rect.get_width() / 2, height / 2, f"{area}\n{due}",
                ha='center', va='bottom', color='black', fontsize=8, rotation=90)


    # Add legend to the chart
    handles = [mpatches.Patch(color=colors[i], label=customer_sales['area'].iloc[i]) for i in range(len(customer_sales))]
    plt.legend(handles=handles, bbox_to_anchor=(1.2, 1), loc='upper left', title='Area Names', fontsize=8)

    # Add x-axis label on the bottom of the chart
    #fig.text(0.5, -0.15, 'Area Names', ha='center', fontsize=10)

    # Add the chart to the tkinter window
    canvas = FigureCanvasTkAgg(fig, master=tab2label0)
    canvas.draw()
    canvas.get_tk_widget().grid(row=1, column=1, padx=0, pady=0, sticky="nsew")
else:
    # Display a message if there is no data available
    label_check1 = customtkinter.CTkLabel(tab2label0, text='No data available to display bar pie',
                                          text_color=('#e5ef56', '#e5ef56'), font=('Arial', 12))
    label_check1.grid(row=1, column=1, padx=10, pady=10)


for widget in tab2label0.winfo_children():
    widget.grid_configure(padx=15, pady=9)


tab2label1 = customtkinter.CTkLabel(master=notebook.tab("Graphs"), text=" ")
tab2label1.grid(row=1, column=0)


df = pd.read_excel(filepath_1)
# Convert payment date to datetime format
df['day'] = pd.to_datetime(df['day'], format='%d-%m-%Y')

# Filter data for current month
current_month1 = datetime.datetime.now().strftime('%B')
active_customers1 = df[df['day'].dt.strftime('%B') == current_month1]['name'].nunique()

# Filter data for customers who paid within 30 days from today's date
today1 = datetime.datetime.now().date()
days_30 = datetime.timedelta(days=30)
paid_customers1 = df[(df['day'].dt.date >= today1 - days_30) & (df['day'].dt.date <= today1)]['name'].unique()

# Filter data for disconnected customers who didn't pay within 30 days from today's date
disconnected_customers1 = df[~df['name'].isin(paid_customers1) & (df['day'].dt.date < today1 - days_30)]['name'].unique()

# Count the number of paid and disconnected customers
paid_count1 = len(paid_customers1)
disconnected_count1 = len(disconnected_customers1)


# Add labels to display the counts of paid and disconnected customers
display_activecustomers = customtkinter.CTkLabel(tab2label1,
                                                 text=f"Active Customers \n(Paid within 30 days):   \n{paid_count1}",
                                                 justify='center', font=("Arial", 13, "bold"),  fg_color=("#e213f4", "#d9f799"),
                                      text_color=("#d9f799","#e213f4"))
display_activecustomers.grid(row=0, column=0, padx=0, pady=10, sticky="n")
display_discustomers = customtkinter.CTkLabel(tab2label1,
                                              text=f"Disconnected Customers \n(Not paid within 30 days):    \n{disconnected_count1}",
                                justify='center', font=("Arial", 13, "bold"),
                                      fg_color=("#cdcff4","#f4cdcd"), text_color=("red","#e213f4"))
display_discustomers.grid(row=0, column=0, padx=10, pady=0, sticky="se")


## Top 10 Due Customers Data


def load_duecustomers_graph():
    df = pd.read_excel(filepath_1)
    # Set the date column as the index
    df.set_index('day', inplace=True)
    # Extract the month and year from the date column
    df['month'] = pd.DatetimeIndex(df.index).month
    df['year'] = pd.DatetimeIndex(df.index).year
    # Group the data by customer and get the sum of sales
    customer_sales2 = df.groupby(['name'])['due'].sum().reset_index()
    # Sort the data by payment in descending order
    customer_sales2 = customer_sales2.sort_values(by='due', ascending=False)
    # Return the top 5 customers by due amount
    return customer_sales2.head(10)

# Load data
customer_sales2 = load_duecustomers_graph()

# Check if there is any data available
if not customer_sales2.empty:
    # Define colors for bars
    colors = ['#1f77b4', '#ff7f0e', '#2ca02c', '#d62728', '#9467bd',
              '#1f77b4', '#ff7f0e', '#2ca02c', '#d62728', '#9467bd']

    # Create bar chart
    fig, ax = plt.subplots(figsize=(4, 2))
    rects1 = ax.bar(customer_sales2['name'], customer_sales2['due'], color=colors[:10])
    ## ax.set_xlabel('Customers Names', fontsize=9)
    #ax.set_ylabel('Due Amount', fontsize=9)
    ax.set_title('Top 10 Customers Due Amount', fontsize=10)
    plt.setp(ax.get_xticklabels(), rotation=0, ha="right", fontsize=8)
    ax.set_xticklabels([])
    ax.set_yticklabels([])
    plt.tight_layout()


    # Add x label names on the bar inside the bar
    for i, v in enumerate(rects1):
        if 'name' in customer_sales2.columns:
            ax.text(i, v.get_height() / 2, f"{customer_sales2['name'].iloc[i]}\n{customer_sales2['due'].iloc[i]}",
                ha='center', fontsize=7, va='bottom', rotation=90, color='black', fontweight='light')
        else:
            ax.text(i, v.get_height() / 2, f"{customer_sales2['name'][i]}\n{customer_sales2['due'][i]}",
                    ha='center', va='bottom', rotation=90, color='black', fontweight='light')


        #if 'name' in customer_sales2.columns:
         #   ax.text(i, v.get_height() / 2, f"{customer_sales2['name'][i]}\n{customer_sales2['due'][i]}", ha='center',
                   # va='bottom', rotation=90, color='black', fontweight='light')
        #else:
         #   ax.text(i, v.get_height() / 2, f"{customer_sales2['name'].iloc[i]}\n{customer_sales2['due'].iloc[i]}",
          #          ha='center', fontsize=7, color='black')

    #for i, v in enumerate(rects1):
       # ax.text(i, v.get_height() / 2, f"{customer_sales2['name'][i]}\n{customer_sales2['due'][i]}", ha='center',
            #    va='bottom', rotation=90, color='black', fontweight='light')

    # Add legend to the chart
    if len(customer_sales2['name']) >= len(colors):
        handles1 = [mpatches.Patch(color=colors[i], label=customer_sales2['name'].iloc[i]) for i in range(len(colors))]
    else:
        handles1 = [mpatches.Patch(color=colors[i], label=customer_sales2['name'].iloc[i]) for i in
                    range(len(customer_sales2['name']))]

    #handles1 = [mpatches.Patch(color=colors[i], label=customer_sales2['name'].iloc[i]) for i in range(len(colors))]
    plt.legend(handles=handles1, bbox_to_anchor=(1.2, 1), loc='upper left', title='Customer Names', fontsize=8)

    # Add x-axis label on the bottom of the chart
    ## fig.text(0.5, -0.15, 'Customer Names', ha='center', fontsize=10)

    # Add the chart to the tkinter window
    canvas1 = FigureCanvasTkAgg(fig, master=tab2label1)
    canvas1.draw()
    canvas1.get_tk_widget().grid(row=0, column=1, padx=0, pady=0, sticky="nsew")
else:
    # Display a message if there is no data available
    label_check2 = customtkinter.CTkLabel(tab2label1, text='No data available to display Customers bar graph',
                                          text_color=('#e5ef56', '#e5ef56'), font=('Arial', 12))
    label_check2.grid(row=0, column=1, padx=10, pady=10)


tab3label3 = customtkinter.CTkLabel(master=notebook.tab("Graphs"), text=" ")
tab3label3.grid(row=2, column=0, padx=342, pady=161)


## Tab 3

def import_data():
    CTkMessagebox(title="Testing stage...", message="Import function is in testing stage, check in next update ",
                  option_1="Thanks")
    print("Error2")
    return


def export_data():
    CTkMessagebox(title="Testing stage...", message="Export function is in testing stage, check in next update ",
                  option_1="Thanks")
    print("Error2")
    return


tab4label0 = customtkinter.CTkLabel(master=notebook.tab("Contact"), text=" ", fg_color=("#333333", "#333333"))
tab4label0.grid(row=0, column=0, padx=10, pady=5)

Contact_frame = customtkinter.CTkLabel(master=tab4label0, text="ॐ",
                                       text_color=("Red", "white"), font=("Arial", 19))
Contact_frame.grid(row=0, column=0)

logo_contants = resource_path("nk32.png")
button_image = customtkinter.CTkImage(Image.open(logo_contants), size=(126, 90))
image_button = customtkinter.CTkButton(master=tab4label0,
                                       image=button_image, text="ॐ \n\n Designed by Niranjan Karupothula "
                                                                "\n S/O. sarangapani \n\n GMail: nirru0007@gmail.com"
                                                                " \n\n All rights reserved \u00A9 ", state=DISABLED,
                                       fg_color=("#333333", "#333333"))
image_button.grid()

label6 = customtkinter.CTkLabel(master=tab4label0, text="Copyright \u00A9 2023", font=("Arial", 13, "italic"),
                                text_color="white")
label6.grid(row=5, column=0, pady=10)


for widget in tab4label0.winfo_children():
    widget.grid_configure(padx=15, pady=9)


tab4label1 = customtkinter.CTkLabel(master=notebook.tab("Contact"), text=" ", fg_color="#333333")
tab4label1.grid(row=1, column=0, padx=10, pady=5)

update_label = customtkinter.CTkLabel(master=tab4label1, text=" About ", font=("Arial", 16, "bold"),
                                      text_color="white")
update_label.grid(row=0, column=1, pady=10)

label2 = customtkinter.CTkLabel(master=tab4label1, text="Customer Data Entry \n\n v1.0", font=("Arial", 15, "italic"),
                                text_color="white")
label2.grid(row=1, column=1, pady=10)


def software_update():
    CTkMessagebox(title="Software Update", message="Software is Up to date!!! ", icon="check",
                  option_1="Thanks")
    print("Software Update Checking")
    return


export_file = customtkinter.CTkButton(master=tab4label1, text="Export Data", command=export_data, width=100)
export_file.grid(row=2, column=0)

update_button = customtkinter.CTkButton(master=tab4label1, text="Software Update", command=software_update)
update_button.grid(row=2, column=1)

import_file = customtkinter.CTkButton(master=tab4label1, text="Import Data", command=import_data, width=100)
import_file.grid(row=2, column=2)

for widget in tab4label1.winfo_children():
    widget.grid_configure(padx=15, pady=9)

tab4label3 = customtkinter.CTkLabel(master=notebook.tab("Contact"), text=" ")
tab4label3.grid(row=2, column=0, padx=342, pady=23)


app.protocol("WM_DELETE_WINDOW", app.quit)

app.mainloop()
